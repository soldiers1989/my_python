# -*- coding:utf-8 -*-



#从数据库查询数据并导入到excle直接打开。

import os
import sys
import time

reload(sys)
from base import DB_connect
from openpyxl import Workbook
print '！！！务必保证所有excle文件现在处于关闭状态！！！'.decode('utf-8')

T=int(time.time()/86400)*86400

sql_member_ship="""SELECT apply_sn,gq_membership_member.member_id,member_name,member_tel,level_id,buy_time,buy_price,binding_type,recommend_user,payment_state,payment_sn,actual_price,arrived_money,member_type,location_pro,location_city
FROM gq_membership_member 
left join tel_location 
on tel_location.tel=member_tel
WHERE payment_state=1 
and member_type=0 
and buy_price>10
AND buy_time<{};
""".format(T)
re= DB_connect.mysql_local_select(sql_member_ship)
#转换成新的时间格式(2016-05-05 20:28:54)
dt = time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(T))
wb = Workbook()
ws = wb.active
name0='会员付费情况'
ws.title=name0.decode('utf-8')
for i in range(0,len(re)):
    ws.cell(row=1, column=1).value ='订单号'
    ws.cell(row=1, column=2).value = '用户id'
    ws.cell(row=1, column=3).value = '用户名称'
    ws.cell(row=1, column=4).value = '手机号'
    ws.cell(row=1, column=5).value = '用户等级'
    ws.cell(row=1, column=6).value = '购买时间（时间戳）'
    ws.cell(row=1, column=7).value = '购买价格'
    ws.cell(row=1, column=8).value = '状态（0正常，1禁用）'
    ws.cell(row=1, column=9).value = '推荐人'
    ws.cell(row=1, column=10).value = '支付状态（1已支付，0未支付）'
    ws.cell(row=1, column=11).value ='支付订单号（支付宝或微信）'
    ws.cell(row=1, column=12).value ='实际价格'
    ws.cell(row=1, column=13).value = '抵扣价格'
    ws.cell(row=1, column=14).value = '用户来源（0正常会员，1后台加入会员）'
    ws.cell(row=1, column=15).value = '省'
    ws.cell(row=1, column=16).value = '市'
    ws.cell(row=1, column=17).value = '购买时间'
    ws.cell(row=1, column=18).value = dt

    buy_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(re[i]['buy_time']))
    ws.cell(row=i + 2, column=1).value = re[i]['apply_sn']
    ws.cell(row=i + 2, column=2).value = re[i]['member_id']
    ws.cell(row=i + 2, column=3).value = re[i]['member_name']
    ws.cell(row=i + 2, column=4).value = re[i]['member_tel']
    ws.cell(row=i + 2, column=5).value = re[i]['level_id']
    ws.cell(row=i + 2, column=6).value = re[i]['buy_time']
    ws.cell(row=i + 2, column=7).value = re[i]['buy_price']
    ws.cell(row=i + 2, column=8).value = re[i]['binding_type']
    ws.cell(row=i + 2, column=9).value = re[i]['recommend_user']
    ws.cell(row=i + 2, column=10).value = re[i]['payment_state']
    ws.cell(row=i + 2, column=11).value = re[i]['payment_sn']
    ws.cell(row=i + 2, column=12).value = re[i]['actual_price']
    ws.cell(row=i + 2, column=13).value = re[i]['arrived_money']
    ws.cell(row=i + 2, column=14).value = re[i]['member_type']
    if re[i]['location_pro']==None:
        ws.cell(row=i + 2, column=15).value = 0
    else:
        ws.cell(row=i + 2, column=15).value = re[i]['location_pro']
    ws.cell(row=i + 2, column=16).value = re[i]['location_city']
    ws.cell(row=i + 2, column=17).value = buy_time

sql_bond="""SELECT member_name,member_tel,external_no,payment_no,payment_code,payment_time,location_pro,location_city  FROM hn_payment_log 
left join hn_member
on bond_payment_no=external_no
left join tel_location
on tel=member_tel
WHERE payment_type=5 and payment_result=1 AND payment_time<{};
""".format(T)
re_bond= DB_connect.mysql_local_select(sql_bond)

ws1 = wb.create_sheet()
name1='保证金缴纳'
ws1.title=name1.decode('utf-8')

for i in range(0,len(re_bond)):
    ws1.cell(row=1, column=1).value ='用户名'
    ws1.cell(row=1, column=2).value = '手机号'
    ws1.cell(row=1, column=3).value = '订单号'
    ws1.cell(row=1, column=4).value = '支付订单号'
    ws1.cell(row=1, column=5).value = '支付类型'
    ws1.cell(row=1, column=6).value = '支付时间（时间戳）'
    ws1.cell(row=1, column=7).value = '省'
    ws1.cell(row=1, column=8).value = '市'
    ws1.cell(row=1, column=9).value = '时间'
    ws1.cell(row=1, column=10).value = dt

    payment_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(re_bond[i]['payment_time']))
    ws1.cell(row=i + 2, column=1).value = re_bond[i]['member_name']
    ws1.cell(row=i + 2, column=2).value = re_bond[i]['member_tel']
    ws1.cell(row=i + 2, column=3).value = re_bond[i]['external_no']
    ws1.cell(row=i + 2, column=4).value = re_bond[i]['payment_no']
    ws1.cell(row=i + 2, column=5).value = re_bond[i]['payment_code']
    ws1.cell(row=i + 2, column=6).value = re_bond[i]['payment_time']
    if re_bond[i]['location_pro']==None:
        ws1.cell(row=i + 2, column=7).value =0
    else:
        ws1.cell(row=i + 2, column=7).value = re_bond[i]['location_pro']
    ws1.cell(row=i + 2, column=8).value = re_bond[i]['location_city']
    ws1.cell(row=i + 2, column=9).value = payment_time
sql_bond_out="""SELECT apply_user_id,apply_user_tel,withdrawals_amount,apply_time,location_pro,location_city
FROM gq_bond_withdrawals
left JOIN tel_location
on apply_user_tel=tel
WHERE audit_status=0;
"""
re_bond_out= DB_connect.mysql_local_select(sql_bond_out)
ws2 = wb.create_sheet()
name2='保证金取消'
ws2.title=name2.decode('utf-8')
for i in range(0,len(re_bond_out)):
    ws2.cell(row=1, column=1).value ='用户id'
    ws2.cell(row=1, column=2).value = '手机号'
    ws2.cell(row=1, column=3).value = '提现金额'
    ws2.cell(row=1, column=4).value = '申请时间戳'
    ws2.cell(row=1, column=5).value = '省'
    ws2.cell(row=1, column=6).value = '市'
    ws2.cell(row=1, column=7).value = '时间'
    ws2.cell(row=1, column=8).value = dt

    apply_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(re_bond_out[i]['apply_time']))
    ws2.cell(row=i + 2, column=1).value = re_bond_out[i]['apply_user_id']
    ws2.cell(row=i + 2, column=2).value = re_bond_out[i]['apply_user_tel']
    ws2.cell(row=i + 2, column=3).value = re_bond_out[i]['withdrawals_amount']
    ws2.cell(row=i + 2, column=4).value = re_bond_out[i]['apply_time']
    if re_bond_out[i]['location_pro']==None:
        ws2.cell(row=i + 2, column=5).value = 0
    else:
        ws2.cell(row=i + 2, column=5).value = re_bond_out[i]['location_pro']
    ws2.cell(row=i + 2, column=6).value = re_bond_out[i]['location_city']
    ws2.cell(row=i + 2, column=7).value = apply_time

wb.save('F:/数据报表-总/付费会员及保证金/付费会员及保证金明细.xlsx'.decode('utf-8'))

print '信息同步完毕，即将打开文档………………'.decode('utf-8')
print '————————————————————————————————————————————'.decode('utf-8')
print '（若文档30秒内未自动打开，直接关闭此程序，手动打开文档或重新运行程序即可。）'.decode('utf-8')

os.startfile('F:/数据报表-总/付费会员及保证金/付费会员及保证金明细.xlsx'.decode('utf-8'))
os.startfile('F:/数据报表-总/付费会员及保证金/每日会员统计_公式.xlsx'.decode('utf-8'))


inp=raw_input("按 回车键 退出:")

print '————————————————————————————————————————————'.decode('utf-8')
print "工作完成，3秒后退出！"
time.sleep(1)

