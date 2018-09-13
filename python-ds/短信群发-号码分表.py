# coding=utf-8


#对excle进行拆分


import xlrd
from openpyxl import Workbook
workbook = xlrd.open_workbook('F:\lings.xlsx')
sheet2 = workbook.sheet_by_name('Sheet1')
rows_need = []
cols = sheet2.col_values(0)  # 获取第一列内容
for i in range(0, len(cols)):
    rows_need.append(sheet2.row_values(i))  # 获取第i 行内容

def excle_save(rows_need,each):
    len_row=int(len(rows_need)/each)
    for j in range(0,len_row):
        name = "phone{}.xlsx".format(j)
        wb = Workbook()
        ws = wb.active
        for i in range(j*each, (j+1)*each):
            ws.cell(row=i-j*each+1, column=1).value = rows_need[i - 1][0]
        wb.save(filename=name)
    wb = Workbook()
    ws = wb.active
    for i in range(len_row * each, len(rows_need)%each+len_row * each):
        ws.cell(row=i - len_row * each + 1, column=1).value = rows_need[i - 1][0]
    wb.save(filename="phone{}.xlsx".format(j+1))
excle_save(rows_need,5000)