#coding: utf-8
from bs4 import BeautifulSoup
import urllib2
import json
import time
from openpyxl import Workbook
import requests
import re

def get_list(member_id,member_pass):
    url = 'https://www.doorock.com/Interface/login?userAccount={}&userPassword={}'.format(member_id, member_pass)
    request = urllib2.Request(url=url,headers={"user-agant":'Mozilla/5.0 (compatible; Baiduspider/2.0;+http://www.baidu.com/search/spider.html)'})
    response = urllib2.urlopen(request)
    soup = BeautifulSoup(response, "html.parser", from_encoding='utf-8')
    res = soup.text
    token = json.loads(res)['token']
    url_get_chat_list='https://www.doorock.com/Interface/chat/searchFriends?token={}&friendName='.format(token)
    headers={
        'Cookie': 'JSESSIONID=7DC51C27058DAD65AE3F14D4AFE62DBF;'
    }
    request = requests.post(url_get_chat_list, headers=headers)
    res_list = request.json()
    re_list = []
    for i in range(1,len(res_list['collection'])):
        list_id=(res_list['collection'][i]['friendId'])
        url_respons = 'https://www.doorock.com/Interface/chat/ajaxChatLog?token={}&chatObject={}'.format(token,
                                                                                                         list_id)
        req = requests.post(url_respons, headers=headers)
        content = req.json()
        chat_respons = content['collection']
        if len(chat_respons)==0:
            pass
        else:
            for i in range(len(chat_respons)):
                if list_id!=chat_respons[i]['senderId']:
                    user_id=('我')
                else:
                    user_id  = chat_respons[i]['senderName']
                content =chat_respons[i]['content']
                picturePic =chat_respons[i]['picturePic']
                createTime =chat_respons[i]['createTime']
                re_list.append([member_id, user_id,content,picturePic,createTime])
    return re_list

member_list=[]

member_list.append(['13001031343','doushi123888'])
member_list.append(['13001199372','doushi123888'])
member_list.append(['13002529157','doushi123888'])
member_list.append(['13002553750','doushi123888'])
member_list.append(['13002554431','doushi123888'])
member_list.append(['13002568968','doushi123888'])
member_list.append(['13002583685','doushi123888'])
member_list.append(['13003342060','doushi123888'])
member_list.append(['13003504086','doushi123888'])
member_list.append(['13003544159','doushi123888'])
member_list.append(['13003555040','doushi123888'])
member_list.append(['13003556328','doushi123888'])
member_list.append(['13004468181','doushi123888'])
member_list.append(['13004474835','doushi123888'])
member_list.append(['13005478763','doushi123888'])
member_list.append(['13005760122','doushi123888'])
member_list.append(['13006666550','doushi123888'])
member_list.append(['13011129560','doushi123888'])
member_list.append(['13011809515','doushi123888'])
member_list.append(['13013708093','doushi123888'])
member_list.append(['13013717662','doushi123888'])
member_list.append(['13013740820','doushi123888'])
member_list.append(['13013778527','doushi123888'])
member_list.append(['13013902171','doushi123888'])
member_list.append(['13013991103','doushi123888'])
member_list.append(['13016711998','doushi123888'])
member_list.append(['13016797768','doushi123888'])
member_list.append(['13016861540','doushi123888'])
member_list.append(['13016922525','doushi123888'])
member_list.append(['13016968141','doushi123888'])
member_list.append(['18118838276','123456789'])
member_list.append(['17366392520','123456789'])
member_list.append(['15295781603','123456789'])
member_list.append(['18021530733','123456789'])
member_list.append(['13182931172','123456789'])
member_list.append(['18395122560','123456789'])
#
member_list.append(['18810983114','123456789'])
#
member_list.append(['15396753319','123456789'])
member_list.append(['18013049934','123456789'])
member_list.append(['15365057814','123456789'])
member_list.append(['13022522360','123456789'])
member_list.append(['13851625374','123456'])
member_list.append(['szqh','123456'])
member_list.append(['gzab','123456'])
member_list.append(['kskl','123456'])
member_list.append(['szqx','123456'])
member_list.append(['szdh','123456'])
member_list.append(['szfk','123456'])
member_list.append(['zhyt','123456'])
member_list.append(['gzdd','123456'])
member_list.append(['jbdz','123456'])
member_list.append(['szxe','123456'])
member_list.append(['13611505733','doushi123888'])
member_list.append(['15850615129','123456789'])
member_list.append(['15025415831','123456789'])
member_list.append(['15295788136','123456789'])
member_list.append(['15295786290','123456789'])
member_list.append(['15951752961','123456789'])
member_list.append(['15295782056','123456789'])
member_list.append(['18114705115','123456789'])
member_list.append(['17898668836','123456789'])
member_list.append(['15295776685','123456789'])
member_list.append(['15295779827','123456789'])
member_list.append(['15852933249','123456789'])
member_list.append(['15295787318','123456789'])
member_list.append(['15077833592','123456789'])
member_list.append(['15387078667','123456789'])
member_list.append(['13401270558','123456789'])
member_list.append(['17314968109','123456789'])
member_list.append(['15950493894','19880610'])
member_list.append(['18724021135','doushi123888'])
member_list.append(['18052028162','doushi123888'])
member_list.append(['18951713976','doushi123888'])
member_list.append(['13917959442','doushi123888'])
member_list.append(['15736026390','doushi123888'])
member_list.append(['17301583192','doushi123888'])
member_list.append(['15298364245','doushi123888'])
#member_list.append(['17374707003','111111liu'])

name = "response.xlsx"
wb = Workbook()
ws = wb.active
data = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
ws.cell(row=1, column=1).value = '账号id'
ws.cell(row=1, column=2).value = '消息发送人名称'
ws.cell(row=1, column=3).value = '消息内容'
ws.cell(row=1, column=4).value = '图片地址'
ws.cell(row=1, column=5).value = '发送时间'
ws.cell(row=1, column=6).value = data
n=0
for i in range(len(member_list)):

    member_id = member_list[i][0]
    member_pass = member_list[i][1]
    list=(get_list(member_id, member_pass))
    if len(list) == 0:
        pass
    else:
        for j in range(len(list)):
            ws.cell(row=n + 2, column=1).value = list[j][0]
            ws.cell(row=n + 2, column=2).value = list[j][1]
            ws.cell(row=n + 2, column=3).value = list[j][2]
            ws.cell(row=n + 2, column=4).value = list[j][3]
            ws.cell(row=n + 2, column=5).value = list[j][4]
            n+=1
            print(list)
wb.save(filename=name)